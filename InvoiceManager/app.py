from __future__ import annotations

import json
import os
import shutil
import sqlite3
import subprocess
import sys
from datetime import date, datetime
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import A5, landscape
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from licensing import install_license, machine_code, read_and_validate_license


APP_NAME = "Finish Invoice Manager"


class ActivationWindow(tk.Tk):
    def __init__(self, reason: str):
        super().__init__()
        self.activated = False
        self.code = machine_code()
        self.title(f"{APP_NAME} - Activation")
        self.geometry("650x360")
        self.resizable(False, False)
        self.configure(bg="#F3F6F9")

        header = tk.Frame(self, bg="#123B4A", height=92)
        header.pack(fill="x")
        header.pack_propagate(False)
        tk.Label(header, text="Software Activation", bg="#123B4A", fg="white",
                 font=("Segoe UI Semibold", 20)).pack(anchor="w", padx=28, pady=(16, 0))
        tk.Label(header, text="This copy must be licensed for this computer.",
                 bg="#123B4A", fg="#C8E1E8", font=("Segoe UI", 10)).pack(anchor="w", padx=28)

        body = tk.Frame(self, bg="#F3F6F9")
        body.pack(fill="both", expand=True, padx=30, pady=22)
        tk.Label(body, text=reason, bg="#F3F6F9", fg="#9B2C2C",
                 font=("Segoe UI", 10), wraplength=580, justify="left").pack(anchor="w")
        tk.Label(body, text="Machine Code", bg="#F3F6F9", fg="#243447",
                 font=("Segoe UI Semibold", 10)).pack(anchor="w", pady=(18, 4))
        entry = ttk.Entry(body, width=58, font=("Consolas", 12))
        entry.insert(0, self.code)
        entry.configure(state="readonly")
        entry.pack(anchor="w")
        buttons = tk.Frame(body, bg="#F3F6F9")
        buttons.pack(anchor="w", pady=20)
        ttk.Button(buttons, text="Copy Machine Code", command=self.copy_code).pack(side="left")
        ttk.Button(buttons, text="Import .lic File", command=self.import_file).pack(side="left", padx=10)
        tk.Label(body, text="Code WhatsApp karein; owner ki bheji hui .lic file yahan import karein.",
                 bg="#F3F6F9", fg="#526773", font=("Segoe UI", 9)).pack(anchor="w")

    def copy_code(self):
        self.clipboard_clear()
        self.clipboard_append(self.code)
        self.update()
        messagebox.showinfo("Copied", "Machine Code copied. Ab WhatsApp par bhej dein.", parent=self)

    def import_file(self):
        path = filedialog.askopenfilename(parent=self, title="Select license file",
                                          filetypes=[("License file", "*.lic"), ("All files", "*.*")])
        if not path:
            return
        valid, message = install_license(path)
        if not valid:
            messagebox.showerror("Activation failed", message, parent=self)
            return
        self.activated = True
        messagebox.showinfo("Activated", message, parent=self)
        self.destroy()


def app_dir() -> Path:
    base = Path(os.getenv("LOCALAPPDATA", Path.home())) / "FinishInvoiceManager"
    base.mkdir(parents=True, exist_ok=True)
    (base / "pdfs").mkdir(exist_ok=True)
    return base


BASE = app_dir()
DB_PATH = BASE / "invoices.db"
SETTINGS_PATH = BASE / "settings.json"


def client_asset_path(*parts) -> Path:
    root = Path(sys.executable).resolve().parent if getattr(sys, "frozen", False) else Path(__file__).resolve().parent
    return root.joinpath(*parts)


DEFAULT_SETTINGS = {
    "company": "Hammad Ashraf Foods",
    "heading": "Finish Invoice",
    "logo": "",
    "accent": "#176B87",
}


def money(value: float) -> str:
    return f"{value:,.2f}" if value % 1 else f"{value:,.0f}"


class Database:
    def __init__(self):
        self.conn = sqlite3.connect(DB_PATH)
        self.conn.execute("""
            CREATE TABLE IF NOT EXISTS invoices (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                invoice_no TEXT NOT NULL,
                invoice_date TEXT NOT NULL,
                distributor TEXT, town TEXT, reference TEXT,
                bilty_no TEXT, transporter TEXT, freight REAL,
                items_json TEXT NOT NULL, created_at TEXT NOT NULL
            )
        """)
        self.conn.execute("""
            CREATE TABLE IF NOT EXISTS products (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL COLLATE NOCASE UNIQUE,
                packing TEXT,
                rate REAL NOT NULL DEFAULT 0,
                updated_at TEXT NOT NULL
            )
        """)
        self.conn.commit()
        if not self.conn.execute("SELECT 1 FROM products LIMIT 1").fetchone():
            for (items_json,) in self.conn.execute("SELECT items_json FROM invoices"):
                try:
                    self.save_products(json.loads(items_json))
                except (TypeError, ValueError, json.JSONDecodeError):
                    continue

    def save(self, data: dict) -> int:
        cur = self.conn.execute("""
            INSERT INTO invoices(invoice_no, invoice_date, distributor, town,
            reference, bilty_no, transporter, freight, items_json, created_at)
            VALUES(?,?,?,?,?,?,?,?,?,?)
        """, (data["invoice_no"], data["invoice_date"], data["distributor"],
              data["town"], data["reference"], data["bilty_no"],
              data["transporter"], data["freight"], json.dumps(data["items"]),
              datetime.now().isoformat(timespec="seconds")))
        self.conn.commit()
        return cur.lastrowid

    def recent(self):
        return self.conn.execute("""
            SELECT id, invoice_no, invoice_date, distributor, town, freight,
                   items_json FROM invoices ORDER BY id DESC LIMIT 300
        """).fetchall()

    def export_backup(self, path):
        destination = sqlite3.connect(path)
        try:
            self.conn.backup(destination)
        finally:
            destination.close()

    def import_backup(self, path):
        source = sqlite3.connect(f"file:{Path(path).resolve().as_posix()}?mode=ro", uri=True)
        try:
            integrity = source.execute("PRAGMA integrity_check").fetchone()
            if not integrity or integrity[0] != "ok":
                raise ValueError("Selected database is damaged or invalid.")
            tables = {row[0] for row in source.execute(
                "SELECT name FROM sqlite_master WHERE type='table'"
            )}
            if not {"invoices", "products"}.issubset(tables):
                raise ValueError("This is not a Finish Invoice Manager database backup.")
            self.conn.commit()
            source.backup(self.conn)
            self.conn.commit()
        finally:
            source.close()

    def get(self, invoice_id: int):
        row = self.conn.execute("SELECT * FROM invoices WHERE id=?", (invoice_id,)).fetchone()
        if not row:
            return None
        keys = [x[1] for x in self.conn.execute("PRAGMA table_info(invoices)")]
        data = dict(zip(keys, row))
        data["items"] = json.loads(data.pop("items_json"))
        return data

    def save_products(self, items):
        now = datetime.now().isoformat(timespec="seconds")
        for item in items:
            name = str(item.get("product", "")).strip()
            if not name:
                continue
            self.conn.execute("""
                INSERT INTO products(name, packing, rate, updated_at)
                VALUES(?,?,?,?)
                ON CONFLICT(name) DO UPDATE SET
                    packing=excluded.packing,
                    rate=excluded.rate,
                    updated_at=excluded.updated_at
            """, (name, str(item.get("packing", "")).strip(),
                  float(item.get("rate", 0) or 0), now))
        self.conn.commit()

    def save_product_names(self, names):
        now = datetime.now().isoformat(timespec="seconds")
        added = 0
        for name in names:
            clean_name = str(name or "").strip()
            if not clean_name:
                continue
            cursor = self.conn.execute("""
                INSERT INTO products(name, packing, rate, updated_at)
                VALUES(?, '', 0, ?)
                ON CONFLICT(name) DO NOTHING
            """, (clean_name, now))
            added += cursor.rowcount
        self.conn.commit()
        return added

    def products(self):
        rows = self.conn.execute("""
            SELECT name, packing, rate FROM products
            ORDER BY name COLLATE NOCASE
        """).fetchall()
        return [{"product": row[0], "packing": row[1] or "", "rate": row[2] or 0}
                for row in rows]


class InvoiceApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_NAME)
        self.minsize(1100, 700)
        self.state("zoomed")
        self.db = Database()
        self.settings = self.load_settings()
        self.last_pdf: Path | None = None
        self.vars = {key: tk.StringVar() for key in (
            "invoice_no", "invoice_date", "distributor", "town", "reference",
            "bilty_no", "transporter", "freight"
        )}
        self.vars["invoice_date"].set(date.today().strftime("%d-%b-%Y"))
        self.vars["freight"].set("0")
        self.setup_style()
        self.build_ui()
        self.new_invoice()

    def load_settings(self):
        settings = DEFAULT_SETTINGS.copy()
        if SETTINGS_PATH.exists():
            try:
                settings.update(json.loads(SETTINGS_PATH.read_text(encoding="utf-8")))
            except Exception:
                pass
        default_logo = client_asset_path("images", "logo.jpeg")
        if not settings.get("logo") and default_logo.exists():
            settings["logo"] = str(default_logo)
        return settings

    def setup_style(self):
        self.configure(bg="#F3F6F9")
        style = ttk.Style(self)
        style.theme_use("clam")
        style.configure("TFrame", background="#F3F6F9")
        style.configure("Card.TFrame", background="white")
        style.configure("TLabel", background="#F3F6F9", foreground="#243447", font=("Segoe UI", 10))
        style.configure("Card.TLabel", background="white", foreground="#263746", font=("Segoe UI", 10))
        style.configure("Title.TLabel", background="#123B4A", foreground="white", font=("Segoe UI Semibold", 20))
        style.configure("Sub.TLabel", background="#123B4A", foreground="#C8E1E8", font=("Segoe UI", 10))
        style.configure("Header.TFrame", background="#123B4A")
        style.configure("TEntry", padding=7, fieldbackground="#F8FAFC", bordercolor="#D9E2EA")
        style.configure("Accent.TButton", font=("Segoe UI Semibold", 10), padding=(16, 9),
                        background=self.settings["accent"], foreground="white")
        style.map("Accent.TButton", background=[("active", "#10566E")])
        style.configure("Secondary.TButton", font=("Segoe UI Semibold", 10), padding=(14, 9),
                        background="#E8F0F4", foreground="#123B4A")
        style.configure("Treeview", rowheight=34, font=("Segoe UI", 10), background="white",
                        fieldbackground="white", borderwidth=0)
        style.configure("Treeview.Heading", font=("Segoe UI Semibold", 10), padding=9,
                        background="#E7EFF3", foreground="#1E3C49")
        style.map("Treeview", background=[("selected", "#CDEAF2")], foreground=[("selected", "#123B4A")])

    def build_ui(self):
        header = tk.Frame(self, bg="#123B4A", height=96)
        header.pack(fill="x")
        header.pack_propagate(False)
        title_box = ttk.Frame(header, style="Header.TFrame")
        title_box.pack(side="left", padx=28, pady=10)
        ttk.Label(title_box, text=self.settings["heading"], style="Title.TLabel").pack(anchor="w")
        ttk.Label(title_box, text=self.settings["company"], style="Sub.TLabel").pack(anchor="w")
        actions = tk.Frame(header, bg="#123B4A")
        actions.pack(side="right", padx=25, pady=17)
        ttk.Button(actions, text="＋ New", style="Secondary.TButton", command=self.new_invoice).pack(side="left", padx=5)
        ttk.Button(actions, text="History", style="Secondary.TButton", command=self.show_history).pack(side="left", padx=5)
        ttk.Button(actions, text="Import Products", style="Secondary.TButton", command=self.import_excel).pack(side="left", padx=5)
        ttk.Button(actions, text="Export Product List", style="Secondary.TButton", command=self.export_excel).pack(side="left", padx=5)
        ttk.Button(actions, text="Settings", style="Secondary.TButton", command=self.show_settings).pack(side="left", padx=5)
        ttk.Button(actions, text="Save PDF", style="Accent.TButton", command=self.create_pdf).pack(side="left", padx=5)
        ttk.Button(actions, text="Print", style="Accent.TButton", command=self.print_invoice).pack(side="left", padx=5)

        outer = ttk.Frame(self)
        outer.pack(fill="both", expand=True, padx=22, pady=16)
        info = ttk.Frame(outer, style="Card.TFrame", padding=16)
        info.pack(fill="x")
        fields = [
            ("Distributor", "distributor", 0, 0, 2), ("Invoice No.", "invoice_no", 0, 2, 1),
            ("Invoice Date", "invoice_date", 0, 3, 1), ("Town", "town", 1, 0, 1),
            ("Reference", "reference", 1, 1, 1), ("Bilty No.", "bilty_no", 1, 2, 1),
            ("Transporter", "transporter", 1, 3, 1),
        ]
        for col in range(4):
            info.columnconfigure(col, weight=1)
        for label, key, row, col, span in fields:
            cell = ttk.Frame(info, style="Card.TFrame")
            cell.grid(row=row, column=col, columnspan=span, sticky="ew", padx=7, pady=5)
            ttk.Label(cell, text=label, style="Card.TLabel").pack(anchor="w")
            ttk.Entry(cell, textvariable=self.vars[key]).pack(fill="x", pady=(3, 0))

        table_card = ttk.Frame(outer, style="Card.TFrame", padding=(16, 12))
        table_card.pack(fill="both", expand=True, pady=(14, 0))
        columns = ("sr", "product", "packing", "qty", "rate", "scheme", "scheme_value", "total")
        self.tree = ttk.Treeview(table_card, columns=columns, show="headings", selectmode="browse")
        headings = ("Sr No.", "Product Name", "Packing", "Quantity", "Rate", "Scheme", "Scheme Value", "Total")
        widths = (55, 350, 150, 90, 110, 85, 115, 125)
        for col, label, width in zip(columns, headings, widths):
            self.tree.heading(col, text=label)
            self.tree.column(col, width=width, minwidth=50, anchor="w" if col in ("product", "packing") else "center")
        self.tree.pack(fill="both", expand=True)
        self.tree.bind("<Double-1>", lambda _e: self.edit_item())
        self.tree.bind("<Delete>", lambda _e: self.delete_item())
        row_actions = ttk.Frame(table_card, style="Card.TFrame")
        row_actions.pack(fill="x", pady=(10, 0))
        ttk.Button(row_actions, text="＋ Add Product", style="Accent.TButton", command=self.add_item).pack(side="left")
        ttk.Button(row_actions, text="Edit Selected", style="Secondary.TButton", command=self.edit_item).pack(side="left", padx=8)
        ttk.Button(row_actions, text="Remove", style="Secondary.TButton", command=self.delete_item).pack(side="left")
        self.summary = ttk.Label(row_actions, text="Total Quantity: 0     Total Amount: 0     Invoice Total: 0",
                                 style="Card.TLabel", font=("Segoe UI Semibold", 12))
        self.summary.pack(side="right", padx=8)
        freight_box = ttk.Frame(row_actions, style="Card.TFrame")
        freight_box.pack(side="right", padx=10)
        ttk.Label(freight_box, text="Freight", style="Card.TLabel").pack(side="left", padx=4)
        ent = ttk.Entry(freight_box, textvariable=self.vars["freight"], width=12)
        ent.pack(side="left")
        ent.bind("<KeyRelease>", lambda _e: self.update_summary())

    def new_invoice(self):
        for key, var in self.vars.items():
            var.set("")
        self.vars["invoice_date"].set(date.today().strftime("%d-%b-%Y"))
        self.vars["freight"].set("0")
        self.vars["invoice_no"].set(str(int(datetime.now().timestamp()))[-6:])
        if hasattr(self, "tree"):
            self.tree.delete(*self.tree.get_children())
            self.update_summary()

    def item_dialog(self, current=None):
        win = tk.Toplevel(self)
        win.title("Product Entry")
        win.transient(self)
        win.grab_set()
        win.configure(bg="white")
        win.resizable(False, False)
        keys = ("product", "packing", "qty", "rate", "scheme", "scheme_value")
        labels = ("Product Name", "Packing", "Quantity", "Rate", "Scheme", "Scheme Value")
        vals = current or ("", "", "1", "0", "0", "0")
        variables = {k: tk.StringVar(value=str(v)) for k, v in zip(keys, vals)}
        product_catalog = self.db.products()
        product_names = [item["product"] for item in product_catalog]
        product_by_name = {item["product"].casefold(): item for item in product_catalog}
        product_entry = None

        def fill_product_details(_event=None):
            selected = product_by_name.get(variables["product"].get().strip().casefold())
            if selected:
                variables["product"].set(selected["product"])
                variables["packing"].set(selected["packing"])
                variables["rate"].set(money(float(selected["rate"])))

        def filter_products(_event=None):
            if not product_entry:
                return
            typed = variables["product"].get().strip().casefold()
            matches = [name for name in product_names if name.casefold().startswith(typed)] if typed else product_names
            product_entry.configure(values=matches)
            if typed and matches:
                product_entry.after_idle(lambda: product_entry.event_generate("<Down>"))

        def calculate_scheme_value(_event=None):
            try:
                qty = float(variables["qty"].get().replace(",", "") or 0)
                scheme = float(variables["scheme"].get().replace(",", "") or 0)
                variables["scheme_value"].set(money(qty * scheme))
            except ValueError:
                pass

        for i, (key, label) in enumerate(zip(keys, labels)):
            ttk.Label(win, text=label, style="Card.TLabel").grid(row=i, column=0, sticky="w", padx=20, pady=8)
            if key == "product":
                entry = ttk.Combobox(win, textvariable=variables[key], width=42,
                                     values=product_names)
                product_entry = entry
                entry.bind("<KeyRelease>", filter_products)
                entry.bind("<<ComboboxSelected>>", fill_product_details)
            else:
                entry = ttk.Entry(win, textvariable=variables[key], width=45)
                if key in ("qty", "scheme"):
                    entry.bind("<KeyRelease>", calculate_scheme_value)
            entry.grid(row=i, column=1, padx=20, pady=8)
            if i == 0:
                entry.focus_set()
        result = []
        def accept():
            try:
                qty = float(variables["qty"].get().replace(",", ""))
                rate = float(variables["rate"].get().replace(",", ""))
                scheme = float(variables["scheme"].get().replace(",", "") or 0)
                scheme_value = float(variables["scheme_value"].get().replace(",", "") or 0)
                if not variables["product"].get().strip():
                    raise ValueError("Product name is required")
                result.extend([variables["product"].get().strip(), variables["packing"].get().strip(),
                               qty, rate, scheme, scheme_value])
                self.db.save_products([{"product": result[0], "packing": result[1], "rate": result[3]}])
                win.destroy()
            except ValueError as exc:
                messagebox.showerror("Invalid entry", str(exc), parent=win)
        buttons = ttk.Frame(win, style="Card.TFrame")
        buttons.grid(row=6, column=0, columnspan=2, sticky="e", padx=20, pady=18)
        ttk.Button(buttons, text="Cancel", style="Secondary.TButton", command=win.destroy).pack(side="left", padx=6)
        ttk.Button(buttons, text="Save Product", style="Accent.TButton", command=accept).pack(side="left")
        win.bind("<Return>", lambda _e: accept())
        self.wait_window(win)
        return result or None

    def add_item(self):
        item = self.item_dialog()
        if item:
            total = item[2] * item[3] - item[5]
            self.tree.insert("", "end", values=(len(self.tree.get_children()) + 1, item[0], item[1],
                             money(item[2]), money(item[3]), money(item[4]), money(item[5]), money(total)))
            self.update_summary()

    def edit_item(self):
        selected = self.tree.selection()
        if not selected:
            return
        values = self.tree.item(selected[0], "values")
        current = (values[1], values[2], *[str(values[i]).replace(",", "") for i in range(3, 7)])
        item = self.item_dialog(current)
        if item:
            total = item[2] * item[3] - item[5]
            self.tree.item(selected[0], values=(values[0], item[0], item[1], money(item[2]), money(item[3]),
                           money(item[4]), money(item[5]), money(total)))
            self.update_summary()

    def delete_item(self):
        selected = self.tree.selection()
        if selected:
            self.tree.delete(selected[0])
            for i, iid in enumerate(self.tree.get_children(), 1):
                vals = list(self.tree.item(iid, "values")); vals[0] = i
                self.tree.item(iid, values=vals)
            self.update_summary()

    def items(self):
        result = []
        for iid in self.tree.get_children():
            v = self.tree.item(iid, "values")
            result.append({"product": v[1], "packing": v[2], "qty": float(str(v[3]).replace(",", "")),
                           "rate": float(str(v[4]).replace(",", "")), "scheme": float(str(v[5]).replace(",", "")),
                           "scheme_value": float(str(v[6]).replace(",", "")), "total": float(str(v[7]).replace(",", ""))})
        return result

    @staticmethod
    def excel_number(value, default=0):
        if value is None or value == "":
            return default
        if isinstance(value, (int, float)):
            return float(value)
        return float(str(value).replace(",", "").strip())

    def export_excel(self):
        default_name = "Product-Names.xlsx"
        path = filedialog.asksaveasfilename(
            parent=self, title="Export product names",
            initialdir=BASE, initialfile=default_name,
            defaultextension=".xlsx", filetypes=[("Excel workbook", "*.xlsx")]
        )
        if not path:
            return
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Products"
            accent = self.settings["accent"].replace("#", "")
            ws["A1"] = "Product Name"
            ws["A1"].font = Font(bold=True, color="FFFFFF")
            ws["A1"].fill = PatternFill("solid", fgColor=accent)
            for row_no, item in enumerate(self.db.products(), 2):
                ws.cell(row_no, 1, item["product"])
            ws.column_dimensions["A"].width = 42
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = f"A1:A{max(2, ws.max_row)}"
            wb.save(path)
            messagebox.showinfo(
                "Excel exported",
                "Simple product list ready. Column A mein sirf product names add karein.",
                parent=self
            )
        except Exception as exc:
            messagebox.showerror("Excel export failed", str(exc), parent=self)

    def import_excel(self):
        path = filedialog.askopenfilename(
            parent=self, title="Import product names",
            filetypes=[("Excel workbook", "*.xlsx"), ("All files", "*.*")]
        )
        if not path:
            return
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            names = []
            seen = set()
            for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
                name = str(row[0] or "").strip()
                if not name or name.casefold() in {"product", "product name", "products"}:
                    continue
                key = name.casefold()
                if key not in seen:
                    seen.add(key)
                    names.append(name)
            if not names:
                raise ValueError("Column A mein koi product name nahi mila.")
            added = self.db.save_product_names(names)
            messagebox.showinfo(
                "Products imported",
                f"{len(names)} names read; {added} new products added.\nAb Add Product mein ye names available hain.",
                parent=self
            )
        except Exception as exc:
            messagebox.showerror("Excel import failed", str(exc), parent=self)

    def totals(self):
        items = self.items()
        qty = sum(x["qty"] for x in items)
        amount = sum(x["total"] for x in items)
        try: freight = float(self.vars["freight"].get().replace(",", "") or 0)
        except ValueError: freight = 0
        return qty, amount, freight, amount - freight

    def update_summary(self):
        qty, amount, freight, invoice_total = self.totals()
        scheme_total = sum(x["scheme_value"] for x in self.items())
        self.summary.configure(text=f"Total Quantity: {money(qty)}     Total Scheme: {money(scheme_total)}     Total Amount: {money(amount)}     Invoice Total: {money(invoice_total)}")

    def collect(self):
        if not self.vars["invoice_no"].get().strip():
            raise ValueError("Invoice number is required.")
        if not self.items():
            raise ValueError("Please add at least one product.")
        data = {k: v.get().strip() for k, v in self.vars.items()}
        try: data["freight"] = float(data["freight"].replace(",", "") or 0)
        except ValueError: raise ValueError("Freight must be a number.")
        data["items"] = self.items()
        self.db.save_products(data["items"])
        return data

    def create_pdf(self, silent=False):
        try:
            data = self.collect()
        except ValueError as exc:
            messagebox.showerror("Cannot create invoice", str(exc))
            return None
        default = BASE / "pdfs" / f"Invoice-{data['invoice_no']}.pdf"
        path = default if silent else filedialog.asksaveasfilename(initialdir=default.parent, initialfile=default.name,
                defaultextension=".pdf", filetypes=[("PDF document", "*.pdf")])
        if not path:
            return None
        path = Path(path)
        self.make_pdf(data, path)
        self.db.save(data)
        self.last_pdf = path
        if not silent:
            messagebox.showinfo("Invoice ready", f"PDF saved successfully:\n{path}")
        return path

    def make_pdf(self, data, path: Path):
        page = landscape(A5)
        doc = SimpleDocTemplate(str(path), pagesize=page, leftMargin=8*mm, rightMargin=8*mm,
                                topMargin=6*mm, bottomMargin=6*mm)
        normal = ParagraphStyle("normal", fontName="Helvetica", fontSize=7.5, leading=9)
        bold = ParagraphStyle("bold", parent=normal, fontName="Helvetica-Bold")
        brand = ParagraphStyle("brand", parent=bold, fontSize=11, leading=13)
        center = ParagraphStyle("center", parent=bold, alignment=TA_CENTER, fontSize=15, leading=18)
        right = ParagraphStyle("right", parent=normal, alignment=TA_RIGHT)
        story = []
        logo = ""
        logo_path = Path(self.settings.get("logo", ""))
        if logo_path.is_file():
            logo = Image(str(logo_path))
            logo._restrictSize(22*mm, 22*mm)
            logo.hAlign = "RIGHT"
        header = Table([[
            Paragraph(f"<b>{self.settings['company']}</b>", brand),
            Paragraph(self.settings["heading"], center),
            logo,
        ]], colWidths=[64.67*mm, 64.66*mm, 64.67*mm], rowHeights=[23*mm])
        header.setStyle(TableStyle([
            ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
            ("ALIGN", (0,0), (0,0), "LEFT"),
            ("ALIGN", (1,0), (1,0), "CENTER"),
            ("ALIGN", (2,0), (2,0), "RIGHT"),
            ("LEFTPADDING", (0,0), (-1,-1), 2),
            ("RIGHTPADDING", (0,0), (-1,-1), 2),
        ]))
        details = Table([
            [Paragraph(f"<b>Distributor:</b> {data['distributor']}", normal),
             Paragraph(f"<b>Invoice No:</b> {data['invoice_no']}<br/><b>Invoice Date:</b> {data['invoice_date']}", normal)],
            [Paragraph(f"<b>Town:</b> {data['town']}<br/><b>REF:</b> {data['reference']}", normal),
             Paragraph(f"<b>Bilty No:</b> {data['bilty_no']}<br/><b>Transporter:</b> {data['transporter']}", normal)]
        ], colWidths=[97*mm, 97*mm])
        details.setStyle(TableStyle([
            ("VALIGN", (0,0), (-1,-1), "TOP"),
            ("ALIGN", (1,0), (1,-1), "RIGHT"),
            ("BOTTOMPADDING", (0,0), (-1,-1), 5),
        ]))
        story.extend([header, details, Spacer(1, 2*mm)])
        rows = [["Sr No.", "Product Name", "Packing", "Quantity", "Rate", "Scheme", "Scheme Value", "Total"]]
        for i, x in enumerate(data["items"], 1):
            rows.append([str(i), x["product"], x["packing"], money(x["qty"]), money(x["rate"]),
                         money(x["scheme"]), money(x["scheme_value"]), money(x["total"])])
        table = Table(rows, repeatRows=1, colWidths=[11*mm, 53*mm, 27*mm, 18*mm, 20*mm, 16*mm, 23*mm, 26*mm])
        table.setStyle(TableStyle([
            ("FONT", (0,0), (-1,0), "Helvetica-Bold", 7), ("FONT", (0,1), (-1,-1), "Helvetica", 7),
            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#E5EEF2")), ("GRID", (0,0), (-1,-1), .45, colors.HexColor("#526773")),
            ("ALIGN", (0,0), (0,-1), "CENTER"), ("ALIGN", (3,1), (-1,-1), "RIGHT"),
            ("VALIGN", (0,0), (-1,-1), "MIDDLE"), ("TOPPADDING", (0,0), (-1,-1), 3), ("BOTTOMPADDING", (0,0), (-1,-1), 3)
        ]))
        story.append(table)
        qty, amount, freight, invoice_total = self.totals()
        scheme_total = sum(x["scheme_value"] for x in data["items"])
        summary = Table([
            [Paragraph(f"<b>Total Quantity:</b> {money(qty)}", right), Paragraph(f"<b>Total Amount:</b> {money(amount)}", right)],
            [Paragraph(f"<b>Total Scheme Value:</b> {money(scheme_total)}", right), ""],
            ["", Paragraph(f"<b>Freight:</b> {money(freight)}", right)],
            ["", Paragraph(f"<b>Invoice Total:</b> {money(invoice_total)}", right)],
        ], colWidths=[97*mm, 97*mm])
        summary.setStyle(TableStyle([("ALIGN", (0,0), (-1,-1), "RIGHT"), ("TOPPADDING", (0,0), (-1,-1), 3)]))
        story.extend([summary, Spacer(1, 8*mm), Paragraph("<b>Date Of Cheque:</b> ____________________", normal),
                      Spacer(1, 9*mm), Paragraph("<b>Signature of Accountant:</b> __________________________", normal)])
        doc.build(story)

    def print_invoice(self):
        path = self.create_pdf(silent=True)
        if not path:
            return
        try:
            if sys.platform == "win32":
                os.startfile(str(path), "print")
            else:
                subprocess.run(["lp", str(path)], check=True)
        except Exception as exc:
            messagebox.showwarning("Print", f"PDF was created, but automatic printing failed.\nOpen and print it manually:\n{path}\n\n{exc}")

    def show_history(self):
        win = tk.Toplevel(self); win.title("Invoice History"); win.geometry("900x560"); win.transient(self)
        tree = ttk.Treeview(win, columns=("id","no","date","distributor","town","total"), show="headings")
        for col, text, width in (("id","ID",60),("no","Invoice No.",110),("date","Date",120),("distributor","Distributor",270),("town","Town",160),("total","Amount",130)):
            tree.heading(col,text=text); tree.column(col,width=width)
        def refresh():
            tree.delete(*tree.get_children())
            for row in self.db.recent():
                total = sum(x["total"] for x in json.loads(row[6]))
                tree.insert("","end",values=(row[0],row[1],row[2],row[3],row[4],money(total)))
        refresh()
        tree.pack(fill="both",expand=True,padx=14,pady=14)
        def load():
            sel=tree.selection()
            if not sel:return
            data=self.db.get(int(tree.item(sel[0],"values")[0])); win.destroy(); self.load_invoice(data)
        def export_database():
            default_name = f"Invoice-History-{datetime.now():%Y-%m-%d}.db"
            path = filedialog.asksaveasfilename(
                parent=win, title="Export invoice history backup", initialdir=BASE,
                initialfile=default_name, defaultextension=".db",
                filetypes=[("Invoice database backup", "*.db")]
            )
            if not path:
                return
            try:
                self.db.export_backup(path)
                messagebox.showinfo("Backup exported", f"History backup saved successfully:\n{path}", parent=win)
            except Exception as exc:
                messagebox.showerror("Backup failed", str(exc), parent=win)
        def import_database():
            path = filedialog.askopenfilename(
                parent=win, title="Import invoice history backup",
                filetypes=[("Invoice database backup", "*.db"), ("All files", "*.*")]
            )
            if not path:
                return
            if not messagebox.askyesno(
                "Restore history backup",
                "Current invoice history and product list will be replaced. Continue?",
                parent=win
            ):
                return
            safety_path = BASE / f"invoices-before-restore-{datetime.now():%Y%m%d-%H%M%S}.db"
            try:
                self.db.export_backup(safety_path)
                self.db.import_backup(path)
                refresh()
                messagebox.showinfo(
                    "Backup imported",
                    f"History restored successfully.\nAutomatic safety backup:\n{safety_path}",
                    parent=win
                )
            except Exception as exc:
                messagebox.showerror("Restore failed", str(exc), parent=win)
        buttons = ttk.Frame(win)
        buttons.pack(pady=(0,14))
        ttk.Button(buttons,text="Load Selected Invoice",style="Accent.TButton",command=load).pack(side="left",padx=5)
        ttk.Button(buttons,text="Export Database Backup",style="Secondary.TButton",command=export_database).pack(side="left",padx=5)
        ttk.Button(buttons,text="Import Database Backup",style="Secondary.TButton",command=import_database).pack(side="left",padx=5)
        tree.bind("<Double-1>",lambda _e:load())

    def load_invoice(self, data):
        for key in self.vars:
            self.vars[key].set(str(data.get(key, "")))
        self.tree.delete(*self.tree.get_children())
        for i,x in enumerate(data["items"],1):
            self.tree.insert("","end",values=(i,x["product"],x["packing"],money(x["qty"]),money(x["rate"]),money(x["scheme"]),money(x["scheme_value"]),money(x["total"])))
        self.update_summary()

    def show_settings(self):
        win=tk.Toplevel(self); win.title("Company Settings"); win.geometry("650x390"); win.transient(self); win.grab_set(); win.configure(bg="white")
        company=tk.StringVar(value=self.settings["company"]); heading=tk.StringVar(value=self.settings["heading"])
        for i,(label,var) in enumerate((("Company Name",company),("Invoice Heading",heading))):
            ttk.Label(win,text=label,style="Card.TLabel").grid(row=i,column=0,padx=20,pady=14,sticky="w")
            ttk.Entry(win,textvariable=var,width=38).grid(row=i,column=1,padx=20,pady=14)
        logo=tk.StringVar(value=self.settings.get("logo", ""))
        ttk.Label(win,text="Invoice Logo",style="Card.TLabel").grid(row=2,column=0,padx=20,pady=14,sticky="w")
        logo_entry=ttk.Entry(win,textvariable=logo,width=48,state="readonly")
        logo_entry.grid(row=2,column=1,padx=(20,5),pady=14,sticky="ew")
        def choose_logo():
            path=filedialog.askopenfilename(
                parent=win,title="Choose invoice logo",
                filetypes=[("Image files","*.png;*.jpg;*.jpeg"),("All files","*.*")]
            )
            if not path:return
            suffix=Path(path).suffix.lower() if Path(path).suffix else ".png"
            saved_logo=BASE / f"company-logo{suffix}"
            shutil.copy2(path,saved_logo)
            logo.set(str(saved_logo))
        ttk.Button(win,text="Choose / Update",style="Secondary.TButton",command=choose_logo).grid(row=2,column=2,padx=(5,20),pady=14)
        ttk.Label(
            win,text="Logo PDF ke top-right corner mein display hoga.",style="Card.TLabel"
        ).grid(row=3,column=1,columnspan=2,padx=20,sticky="w")
        def save():
            self.settings.update(company=company.get().strip(),heading=heading.get().strip(),logo=logo.get().strip())
            SETTINGS_PATH.write_text(json.dumps(self.settings,indent=2),encoding="utf-8")
            messagebox.showinfo("Settings","Settings saved. New PDFs will use the updated logo and heading.",parent=win); win.destroy()
        win.columnconfigure(1,weight=1)
        ttk.Button(win,text="Save Settings",style="Accent.TButton",command=save).grid(row=4,column=0,columnspan=3,pady=34)


if __name__ == "__main__":
    valid, reason, _license = read_and_validate_license()
    if not valid:
        activation = ActivationWindow(reason)
        activation.mainloop()
        valid = activation.activated
    if valid:
        InvoiceApp().mainloop()
